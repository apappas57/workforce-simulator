"""utils/excel_export.py

Formatted Excel (.xlsx) workbook builder for the Workforce Simulator.

Produces a multi-sheet workbook with consistent styling:
  • Bold header row with indigo fill and white text
  • Freeze pane on row 1
  • Auto-sized column widths (capped at 40)
  • Number formatting per column type
  • Alternating row shading (light zinc)
  • A Summary sheet with headline KPIs

Public API
----------
build_simulation_workbook(
    df_inputs, df_erlang, roster_df=None,
    planning_df=None, optimisation_df=None,
    cost_interval_df=None, cost_monthly_df=None,
    des_daily_df=None,
) -> bytes          (the .xlsx file as raw bytes)
"""

from __future__ import annotations

import io
from typing import Optional

import pandas as pd

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
_HEADER_FILL   = "4F46E5"   # indigo-600
_ALT_ROW_FILL  = "F4F4F5"   # zinc-100
_BORDER_COLOUR = "D4D4D8"   # zinc-300
_WHITE         = "FFFFFF"
_TEXT_DARK     = "18181B"    # zinc-900

_MAX_COL_WIDTH = 40
_MIN_COL_WIDTH = 8


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def _header_fill():
    return PatternFill("solid", fgColor=_HEADER_FILL)

def _alt_fill():
    return PatternFill("solid", fgColor=_ALT_ROW_FILL)

def _thin_border():
    side = Side(style="thin", color=_BORDER_COLOUR)
    return Border(left=side, right=side, top=side, bottom=side)

def _header_font():
    return Font(bold=True, color=_WHITE, name="Calibri", size=10)

def _body_font():
    return Font(color=_TEXT_DARK, name="Calibri", size=10)

def _centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _left():
    return Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Column format detection
# ---------------------------------------------------------------------------

_PCT_HINTS   = {"sl", "sla", "occ", "occupancy", "shrink", "attrition", "pct", "%",
                "rate", "fraction", "utilisation"}
_MONEY_HINTS = {"cost", "spend", "labour", "penalty", "idle", "saving"}
_INT_HINTS   = {"agents", "staff", "fte", "headcount", "hc", "calls", "interval",
                "count", "n_", "num_", "hire", "surplus", "deficit", "gap"}


def _col_format(col_name: str, dtype) -> Optional[str]:
    """Return an Excel number-format string for *col_name*."""
    cn = col_name.lower()

    if pd.api.types.is_datetime64_any_dtype(dtype):
        return "YYYY-MM-DD HH:MM"

    if pd.api.types.is_float_dtype(dtype):
        if any(h in cn for h in _PCT_HINTS):
            return "0.0%"
        if any(h in cn for h in _MONEY_HINTS):
            return '#,##0.00'
        if any(h in cn for h in _INT_HINTS):
            return '#,##0'
        return '#,##0.00'

    if pd.api.types.is_integer_dtype(dtype):
        return '#,##0'

    return None   # General (string, object, etc.)


# ---------------------------------------------------------------------------
# Core sheet writer
# ---------------------------------------------------------------------------

def _write_sheet(
    ws,
    df: pd.DataFrame,
    sheet_title: str,
    pct_columns: Optional[set] = None,
) -> None:
    """Write *df* to worksheet *ws* with full formatting."""

    pct_columns = pct_columns or set()

    headers = list(df.columns)

    # ── Header row ──────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = _header_fill()
        cell.font   = _header_font()
        cell.border = _thin_border()
        cell.alignment = _centre()

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    # ── Data rows ────────────────────────────────────────────────────────────
    col_fmts = {
        col: _col_format(col, df[col].dtype)
        for col in headers
    }

    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        use_alt = (row_idx % 2 == 0)
        row_fill = _alt_fill() if use_alt else None

        for col_idx, (col_name, value) in enumerate(zip(headers, row_data), start=1):
            # Coerce pandas NA / NaT to None so Excel renders blank
            if pd.isna(value) if not isinstance(value, str) else False:
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = _body_font()
            cell.border = _thin_border()
            cell.alignment = _left()
            if row_fill:
                cell.fill = row_fill

            fmt = col_fmts.get(col_name)
            if fmt:
                cell.number_format = fmt

    # ── Column widths ────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        sample_vals = df[col_name].dropna().astype(str).head(200)
        max_len = max(
            (len(str(v)) for v in sample_vals),
            default=_MIN_COL_WIDTH,
        )
        width = min(max(max_len + 2, len(col_name) + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _write_summary(ws, kpis: list[tuple[str, str]]) -> None:
    """Write a two-column KPI table (Metric, Value)."""

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18

    headers = ["Metric", "Value"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = _header_fill()
        cell.font      = _header_font()
        cell.border    = _thin_border()
        cell.alignment = _centre()

    ws.freeze_panes = "A2"

    for row_idx, (metric, value) in enumerate(kpis, start=2):
        for col_idx, v in enumerate([metric, value], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font      = _body_font()
            cell.border    = _thin_border()
            cell.alignment = _left()
            if row_idx % 2 == 0:
                cell.fill = _alt_fill()


# ---------------------------------------------------------------------------
# KPI extraction helpers
# ---------------------------------------------------------------------------

def _erlang_kpis(df: pd.DataFrame) -> list[tuple[str, str]]:
    kpis = []
    try:
        kpis.append(("Total calls offered",         f"{df['calls_offered'].sum():,.0f}"))
        kpis.append(("Peak agents required",         f"{df['agents_required'].max():,.0f}"))
        kpis.append(("Avg agents required",          f"{df['agents_required'].mean():,.1f}"))
        if "erlang_sl_pct" in df.columns:
            kpis.append(("Avg SL % (Erlang C)",      f"{df['erlang_sl_pct'].mean():.1%}"))
        if "erlang_occupancy" in df.columns:
            kpis.append(("Avg occupancy % (Erlang C)", f"{df['erlang_occupancy'].mean():.1%}"))
        if "aht_seconds" in df.columns:
            kpis.append(("AHT (seconds)",            f"{df['aht_seconds'].mean():,.0f}"))
    except Exception:
        pass
    return kpis


def _planning_kpis(df: pd.DataFrame) -> list[tuple[str, str]]:
    kpis = []
    try:
        kpis.append(("Planning horizon (months)", f"{len(df):,}"))
        kpis.append(("Opening headcount",          f"{df['opening_hc'].iloc[0]:,.0f}"))
        kpis.append(("Closing headcount",           f"{df['closing_hc'].iloc[-1]:,.0f}"))
        if "capacity_gap" in df.columns:
            deficit_months = (df["capacity_gap"] < 0).sum()
            kpis.append(("Months in deficit",       f"{deficit_months}"))
        if "total_hires" in df.columns:
            kpis.append(("Total hires planned",     f"{df['total_hires'].sum():,.0f}"))
    except Exception:
        pass
    return kpis


def _cost_kpis(df: pd.DataFrame) -> list[tuple[str, str]]:
    kpis = []
    try:
        kpis.append(("Total labour cost",  f"${df['labour_cost'].sum():,.2f}"))
        kpis.append(("Total idle cost",    f"${df['idle_cost'].sum():,.2f}"))
        kpis.append(("Total breach cost",  f"${df['breach_cost'].sum():,.2f}"))
        total = df["labour_cost"].sum() + df["idle_cost"].sum() + df["breach_cost"].sum()
        kpis.append(("Grand total cost",   f"${total:,.2f}"))
        if "calls_offered" in df.columns and df["calls_offered"].sum() > 0:
            cpc = total / df["calls_offered"].sum()
            kpis.append(("Cost per call",  f"${cpc:,.4f}"))
    except Exception:
        pass
    return kpis


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_simulation_workbook(
    df_inputs:         pd.DataFrame,
    df_erlang:         pd.DataFrame,
    roster_df:         Optional[pd.DataFrame] = None,
    planning_df:       Optional[pd.DataFrame] = None,
    optimisation_df:   Optional[pd.DataFrame] = None,
    cost_interval_df:  Optional[pd.DataFrame] = None,
    cost_monthly_df:   Optional[pd.DataFrame] = None,
    des_daily_df:      Optional[pd.DataFrame] = None,
) -> bytes:
    """Build and return a formatted .xlsx workbook as raw bytes.

    Raises RuntimeError if openpyxl is not installed.
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError(
            "openpyxl is required for Excel export. "
            "Run: pip install openpyxl"
        )

    wb = Workbook()

    # ── Summary ───────────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    all_kpis: list[tuple[str, str]] = []
    all_kpis += _erlang_kpis(df_erlang)
    if planning_df is not None and not planning_df.empty:
        all_kpis += _planning_kpis(planning_df)
    if cost_interval_df is not None and not cost_interval_df.empty:
        all_kpis += _cost_kpis(cost_interval_df)
    _write_summary(ws_summary, all_kpis)

    # ── Demand ────────────────────────────────────────────────────────────────
    ws_demand = wb.create_sheet("Demand")
    _write_sheet(ws_demand, df_inputs.copy(), "Demand")

    # ── Erlang C results ──────────────────────────────────────────────────────
    ws_erlang = wb.create_sheet("Erlang C")
    erlang_export = df_erlang.copy()
    # Coerce timestamp columns to plain strings for Excel compat
    for col in erlang_export.select_dtypes(include=["datetimetz"]).columns:
        erlang_export[col] = erlang_export[col].dt.strftime("%Y-%m-%d %H:%M")
    _write_sheet(ws_erlang, erlang_export, "Erlang C")

    # ── Roster ────────────────────────────────────────────────────────────────
    if roster_df is not None and not roster_df.empty:
        ws_roster = wb.create_sheet("Roster")
        roster_export = roster_df.copy()
        for col in roster_export.select_dtypes(include=["datetimetz"]).columns:
            roster_export[col] = roster_export[col].dt.strftime("%Y-%m-%d %H:%M")
        _write_sheet(ws_roster, roster_export, "Roster")

    # ── DES daily summary ─────────────────────────────────────────────────────
    if des_daily_df is not None and not des_daily_df.empty:
        ws_des = wb.create_sheet("Simulation")
        _write_sheet(ws_des, des_daily_df.copy(), "Simulation")

    # ── Workforce planning ────────────────────────────────────────────────────
    if planning_df is not None and not planning_df.empty:
        ws_plan = wb.create_sheet("Planning")
        plan_export = planning_df.copy()
        if "period_start" in plan_export.columns:
            plan_export["period_start"] = plan_export["period_start"].dt.strftime("%Y-%m-%d")
        _write_sheet(ws_plan, plan_export, "Planning")

    # ── Hiring optimisation ───────────────────────────────────────────────────
    if optimisation_df is not None and not optimisation_df.empty:
        ws_opt = wb.create_sheet("Optimisation")
        opt_export = optimisation_df.copy()
        if "period_start" in opt_export.columns:
            opt_export["period_start"] = opt_export["period_start"].dt.strftime("%Y-%m-%d")
        _write_sheet(ws_opt, opt_export, "Optimisation")

    # ── Cost — interval ───────────────────────────────────────────────────────
    if cost_interval_df is not None and not cost_interval_df.empty:
        ws_cost = wb.create_sheet("Cost — Interval")
        cost_iv_export = cost_interval_df.copy()
        for col in cost_iv_export.select_dtypes(include=["datetimetz"]).columns:
            cost_iv_export[col] = cost_iv_export[col].dt.strftime("%Y-%m-%d %H:%M")
        _write_sheet(ws_cost, cost_iv_export, "Cost — Interval")

    # ── Cost — monthly projection ─────────────────────────────────────────────
    if cost_monthly_df is not None and not cost_monthly_df.empty:
        ws_cost_m = wb.create_sheet("Cost — Monthly")
        cost_m_export = cost_monthly_df.copy()
        if "period_start" in cost_m_export.columns:
            cost_m_export["period_start"] = cost_m_export["period_start"].dt.strftime("%Y-%m-%d")
        _write_sheet(ws_cost_m, cost_m_export, "Cost — Monthly")

    # ── Serialise to bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
