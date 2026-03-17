"""tests/test_excel_export.py

Unit tests for utils/excel_export.py — Phase 17 formatted Excel export.

Tests run only when openpyxl is installed (which it is in the standard
requirements.txt install).  The fallback RuntimeError for a missing openpyxl
is also tested via mocking.

Covers:
  - build_simulation_workbook() returns valid bytes
  - Resulting workbook contains expected sheet names
  - Optional DataFrames (roster, planning, cost, DES) don't break the build
  - Empty erlang_df handled gracefully
  - RuntimeError raised when _OPENPYXL_AVAILABLE is False
"""

import io
import unittest
from unittest.mock import patch

import pandas as pd

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from utils.excel_export import build_simulation_workbook


# ---------------------------------------------------------------------------
# Minimal test fixtures
# ---------------------------------------------------------------------------

def _demand_df():
    return pd.DataFrame({
        "interval":          [0, 1, 2, 3],
        "calls_offered":     [50.0, 60.0, 70.0, 80.0],
        "aht_seconds_used":  [240, 240, 240, 240],
    })


def _erlang_df():
    return pd.DataFrame({
        "interval":                       [0, 1, 2, 3],
        "calls_offered":                  [50.0, 60.0, 70.0, 80.0],
        "erlang_required_net_agents":     [5.0, 6.0, 7.0, 8.0],
        "erlang_required_paid_agents":    [6.0, 7.0, 8.0, 9.0],
        "erlang_pred_service_level":      [0.85, 0.87, 0.88, 0.90],
        "erlang_pred_occupancy":          [0.72, 0.74, 0.76, 0.78],
    })


def _roster_df():
    return pd.DataFrame({
        "interval":          [0, 1, 2, 3],
        "roster_net_agents": [6.0, 7.0, 8.0, 9.0],
    })


def _planning_df():
    return pd.DataFrame({
        "month":         [1, 2, 3],
        "available_fte": [50.0, 52.0, 55.0],
        "required_fte":  [48.0, 51.0, 53.0],
    })


def _cost_interval_df():
    return pd.DataFrame({
        "interval":        [0, 1, 2, 3],
        "labour_cost":     [75.0, 90.0, 105.0, 120.0],
        "total_cost":      [80.0, 96.0, 111.0, 127.0],
        "calls_offered":   [50.0, 60.0, 70.0, 80.0],
    })


# ---------------------------------------------------------------------------
# Tests (skipped if openpyxl absent)
# ---------------------------------------------------------------------------

@unittest.skipUnless(_OPENPYXL_AVAILABLE, "openpyxl not installed — skipping Excel export tests")
class TestBuildSimulationWorkbook(unittest.TestCase):

    def test_returns_bytes(self):
        result = build_simulation_workbook(_demand_df(), _erlang_df())
        self.assertIsInstance(result, bytes)

    def test_bytes_are_non_empty(self):
        result = build_simulation_workbook(_demand_df(), _erlang_df())
        self.assertGreater(len(result), 0)

    def test_result_is_valid_xlsx(self):
        result = build_simulation_workbook(_demand_df(), _erlang_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        # If we reach here, the bytes represent a valid workbook
        self.assertGreater(len(wb.sheetnames), 0)

    def test_summary_sheet_always_present(self):
        result = build_simulation_workbook(_demand_df(), _erlang_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        self.assertIn("Summary", wb.sheetnames)

    def test_erlang_sheet_present(self):
        result = build_simulation_workbook(_demand_df(), _erlang_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        sheet_names_lower = [s.lower() for s in wb.sheetnames]
        self.assertTrue(
            any("erlang" in s for s in sheet_names_lower),
            f"No 'Erlang' sheet found in {wb.sheetnames}",
        )

    def test_with_all_optional_dataframes(self):
        """All optional args provided — should produce a valid workbook."""
        result = build_simulation_workbook(
            _demand_df(),
            _erlang_df(),
            roster_df=_roster_df(),
            planning_df=_planning_df(),
            cost_interval_df=_cost_interval_df(),
        )
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_empty_erlang_df_does_not_raise(self):
        """Empty Erlang DataFrame should not crash the builder."""
        try:
            result = build_simulation_workbook(_demand_df(), pd.DataFrame())
            self.assertIsInstance(result, bytes)
        except Exception as exc:
            self.fail(f"build_simulation_workbook raised unexpectedly: {exc}")


# ---------------------------------------------------------------------------
# Fallback behaviour when openpyxl is absent
# ---------------------------------------------------------------------------

class TestOpenpyxlFallback(unittest.TestCase):

    def test_raises_runtime_error_when_openpyxl_missing(self):
        """RuntimeError raised with a helpful install message when openpyxl absent."""
        with patch("utils.excel_export._OPENPYXL_AVAILABLE", False):
            with self.assertRaises(RuntimeError) as ctx:
                build_simulation_workbook(_demand_df(), _erlang_df())
        self.assertIn("openpyxl", str(ctx.exception).lower())


if __name__ == "__main__":
    unittest.main()
